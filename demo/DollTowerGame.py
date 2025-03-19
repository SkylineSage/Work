import random
import pandas as pd
import gradio as gr
from collections import defaultdict
import tempfile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

DEFAULT_COLORS = ["红", "橙", "黄", "绿", "蓝", "紫", "粉", "黑", "白", "棕", "灰", "金", "银"]


def parse_doll_config(config_str):
    """解析娃娃配置字符串"""
    colors = DEFAULT_COLORS.copy()
    ratios = []
    for i, color in enumerate(colors):
        if i < len(config_str.split(",")):
            part = config_str.split(",")[i].strip()
            ratio = 1.0
            try:
                ratio = float(part)
                if ratio <= 0:
                    ratio = 0.0
            except ValueError:
                ratio = 0.0
        else:
            ratio = 0.0
        ratios.append(ratio)

    # 检查并处理比例
    sum_ratios = sum(r for r in ratios if r > 0)
    if sum_ratios == 0:
        sum_ratios = 1.0
    for i in range(len(ratios)):
        if ratios[i] <= 0:
            ratios[i] = 0.0
        else:
            ratios[i] /= sum_ratios
    return colors, ratios


class GameState:
    """游戏状态管理类"""

    def __init__(self, game_id, population, ratios, initial_draw, wish_colors, milk_counts):
        self.game_id = game_id
        self.population = population
        self.ratios = ratios
        self.wish_colors = wish_colors  # 许愿色改为列表
        self.milk_counts = milk_counts  # 新增奶次数
        self.milk_used = 0  # 已使用的奶次数

        # 初始化状态
        self.tower = [None] * 9  # 9层宝塔
        self.small_basket = []
        self.harvest_dolls = 0
        self.harvest_gifts = 0
        self.round_number = 0
        self.logs = []

        # 初始抽取
        initial = random.choices(population, weights=ratios, k=initial_draw)
        for i in range(min(9, len(initial))):
            self.tower[i] = {"color": initial[i], "wish_triggered": False}
        self.small_basket = initial[9:]

        # 初始日志记录
        self.log_event("初始化", f"初始放置 {len(initial)} 个娃娃")

    def log_event(self, phase, event):
        """记录游戏事件"""
        tower_count = sum(1 for d in self.tower if d)
        self.logs.append([
            self.game_id,
            self.round_number,
            phase,
            self.tower_str(),
            len(self.small_basket),
            self.harvest_dolls,
            self.harvest_gifts,
            tower_count,
            event
        ])

    def tower_str(self):
        """宝塔状态可视化"""
        return " | ".join(
            f"{d['color']: <3}{'*' if d['wish_triggered'] else ' '}" if d else "●   "
            for d in self.tower
        )

    def process_wish(self):
        """处理许愿色补货（支持多选）"""
        wish_count = 0
        for color in self.wish_colors:
            wish_count += sum(
                1 for d in self.tower
                if d and d["color"] == color and not d["wish_triggered"]
            )

        if wish_count > 0:
            # 标记已触发
            for d in self.tower:
                if d and d["color"] in self.wish_colors:
                    d["wish_triggered"] = True

            # 补货到小筐
            new_dolls = random.choices(self.population, weights=self.ratios, k=wish_count)
            self.small_basket.extend(new_dolls)
            self.log_event("许愿补货", f"补充 {wish_count} 个娃娃")

    def process_groups(self):
        """处理同色组"""
        groups = [(0, 3, 3), (3, 5, 2), (5, 7, 2), (7, 9, 2)]

        for idx, (start, end, req) in enumerate(groups, 1):
            group = self.tower[start:end]

            # 检查是否填满且同色
            if None in group:
                continue

            colors = set(d["color"] for d in group)
            if len(colors) == 1:
                # 收获娃娃
                self.harvest_dolls += (end - start)

                # 清空组
                for i in range(start, end):
                    self.tower[i] = None

                # 补货到小筐
                new_dolls = random.choices(self.population, weights=self.ratios, k=end - start)
                self.small_basket.extend(new_dolls)

                self.log_event("组处理", f"第{idx}组清除 {end - start} 个，补货 {len(new_dolls)} 个")

    def process_duplicates(self):
        """处理全塔重复颜色（修复补货逻辑）"""
        color_counts = defaultdict(int)
        for doll in self.tower:
            if doll:
                color_counts[doll["color"]] += 1

        # 计算需要移除的颜色和补货数量
        removed_colors = {color for color, count in color_counts.items() if count >= 2}
        total_replenish = sum((count - 1) for color, count in color_counts.items() if count >= 2)

        # 移除娃娃
        removed = 0
        for i in range(len(self.tower)):
            if self.tower[i] and self.tower[i]["color"] in removed_colors:
                self.tower[i] = None
                removed += 1

        if removed > 0:
            self.harvest_dolls += removed
            # 按规则补货：每个颜色组补货n-1
            new_dolls = random.choices(self.population, weights=self.ratios, k=total_replenish)
            self.small_basket.extend(new_dolls)
            self.log_event("同色处理", f"清除 {removed} 个重复颜色，补货 {total_replenish} 个")

    def process_special_gift(self):
        """处理特殊大礼包"""
        dolls = [d for d in self.tower if d]
        if len(dolls) == 9 and len(set(d["color"] for d in dolls)) == 9:
            # 获得大礼包
            self.harvest_gifts += 1
            self.log_event("大礼包", "获得全异色大礼包")

            # 将宝塔上的所有娃娃倒入收获筐
            self.harvest_dolls += 9
            self.tower = [None] * 9  # 清空宝塔
            self.log_event("大礼包", "宝塔清空，所有娃娃倒入收获筐")

    def refill_tower(self):
        """补充娃娃到宝塔"""
        if not self.small_basket:
            return

        # 打乱顺序后填充
        random.shuffle(self.small_basket)
        new_dolls = iter(self.small_basket)

        # 填充空位
        filled = 0
        for i in range(9):
            if self.tower[i] is None:
                try:
                    color = next(new_dolls)
                    self.tower[i] = {"color": color, "wish_triggered": False}
                    filled += 1
                except StopIteration:
                    break

        # 更新小筐
        self.small_basket = list(new_dolls)
        if filled > 0:
            self.log_event("新一轮放娃", f"放置 {filled} 个娃娃到宝塔")

    def should_terminate(self):
        """检查终止条件（支持多选许愿色）"""
        # 检查小筐是否为空
        if len(self.small_basket) > 0:
            return False

        # 检查许愿色是否全部触发
        if any(
                d for d in self.tower
                if d and d["color"] in self.wish_colors and not d["wish_triggered"]
        ):
            return False

        # 检查是否有重复颜色
        colors = [d["color"] for d in self.tower if d]
        return len(colors) == len(set(colors))

    def apply_milk(self):
        """应用奶操作"""
        if self.milk_used >= len(self.milk_counts):
            return False

        milk_count = self.milk_counts[self.milk_used]
        if milk_count <= 0:
            return False

        # 从population中随机抽取娃娃
        new_dolls = random.choices(self.population, weights=self.ratios, k=milk_count)

        # 放置到宝塔
        filled = 0
        for i in range(9):
            if self.tower[i] is None:
                self.tower[i] = {"color": new_dolls[filled], "wish_triggered": False}
                filled += 1
                if filled >= milk_count:
                    break

        # 剩余的放入小筐
        self.small_basket.extend(new_dolls[filled:])

        self.milk_used += 1
        self.log_event("奶操作", f"应用第{self.milk_used}次奶，补充{milk_count}个娃娃")
        return True


def simulate_game(game_id, config, max_rounds=100):
    """单局游戏模拟"""
    state = GameState(
        game_id=game_id,
        population=config["population"],
        ratios=config["ratios"],
        initial_draw=config["initial_draw"],
        wish_colors=config["wish_colors"],
        milk_counts=config["milk_counts"]
    )

    while state.round_number < max_rounds:
        state.round_number += 1

        # 执行处理步骤
        state.process_wish()
        state.process_groups()
        state.process_duplicates()
        state.process_special_gift()

        # 补货阶段
        state.refill_tower()

        # 检查终止条件
        if state.should_terminate():
            # 尝试应用奶操作
            if state.apply_milk():
                continue

            # 游戏结束时清空宝塔
            remaining = sum(1 for d in state.tower if d)
            if remaining > 0:
                state.harvest_dolls += remaining
                state.tower = [None] * 9
                state.log_event("游戏结束", f"清空宝塔，收获{remaining}个娃娃，正常终止")
            break
    else:
        # 达到最大回合时清空宝塔
        remaining = sum(1 for d in state.tower if d)
        if remaining > 0:
            state.harvest_dolls += remaining
            state.tower = [None] * 9
            state.log_event("游戏结束", f"达到最大回合，清空宝塔收获{remaining}个娃娃")

    # 最终礼包兑换
    exchange = state.harvest_dolls // config["doll_exchange"]
    state.harvest_gifts += exchange
    final_dolls = state.harvest_dolls % config["doll_exchange"]

    return {
        "dolls": final_dolls,
        "gifts": state.harvest_gifts,
        "rounds": state.round_number
    }, state.logs


def run_simulation(doll_config_red, doll_config_orange, doll_config_yellow, doll_config_green, doll_config_blue,
                   doll_config_purple, doll_config_pink, doll_config_black, doll_config_white, doll_config_brown,
                   doll_config_gray, doll_config_gold, doll_config_silver, wish_colors, initial_draw, doll_exchange,
                   total_games, max_rounds, milk1, milk2, milk3):
    """批量运行模拟（支持多选许愿色）"""
    try:
        # 将各个颜色的数值拼接成字符串，传递给 parse_doll_config
        doll_config = f"{doll_config_red},{doll_config_orange},{doll_config_yellow},{doll_config_green},{doll_config_blue},{doll_config_purple},{doll_config_pink},{doll_config_black},{doll_config_white},{doll_config_brown},{doll_config_gray},{doll_config_gold},{doll_config_silver}"
        population, ratios = parse_doll_config(doll_config)
        config = {
            "population": population,
            "ratios": ratios,
            "initial_draw": int(initial_draw),
            "wish_colors": wish_colors,
            "doll_exchange": int(doll_exchange),
            "milk_counts": [int(milk1), int(milk2), int(milk3)]  # 新增奶参数
        }

        # 检查许愿色是否在娃娃颜色列表中
        for color in wish_colors:
            if color not in population:
                return f"错误：许愿色 {color} 不在娃娃颜色列表中！", None

        results = []
        all_logs = []

        for game_id in range(1, int(total_games) + 1):
            result, logs = simulate_game(game_id, config, max_rounds)
            results.append([game_id, result["dolls"], result["gifts"], result["rounds"]])
            all_logs.extend(logs)

        # 生成数据文件
        df_logs = pd.DataFrame(all_logs, columns=[
            "游戏ID", "回合数", "阶段", "宝塔状态", "小筐数量",
            "收获娃娃", "大礼包数", "塔上娃娃", "事件描述"
        ])
        df_results = pd.DataFrame(results, columns=[
            "游戏ID", "收获娃娃", "大礼包数", "总回合数"
        ])

        # 新增参数表
        df_params = pd.DataFrame({
            "参数": ["红色", "橙色", "黄色", "绿色", "蓝色", "紫色", "粉色", "黑色", "白色", "棕色", "灰色", "金色",
                     "银色", "许愿色", "初始抽取", "娃娃兑换率", "总局数", "第一次奶", "第二次奶", "第三次奶"],
            "值": [doll_config_red, doll_config_orange, doll_config_yellow, doll_config_green, doll_config_blue,
                   doll_config_purple, doll_config_pink, doll_config_black, doll_config_white, doll_config_brown,
                   doll_config_gray, doll_config_gold, doll_config_silver, ", ".join(wish_colors), initial_draw, doll_exchange, total_games,
                   milk1, milk2, milk3]
        })

        # 将数据保存为临时文件
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
                df_logs.to_excel(writer, sheet_name="详细日志", index=False)
                df_results.to_excel(writer, sheet_name="汇总结果", index=False)
                df_params.to_excel(writer, sheet_name="参数设置", index=False)  # 新增参数表

            # 加载Excel文件以设置样式
            wb = load_workbook(tmp.name)
            ws = wb["详细日志"]

            # 设置“宝塔状态”列的列宽
            column_letter = get_column_letter(df_logs.columns.get_loc("宝塔状态") + 1)  # +1因为Excel列索引从1开始
            ws.column_dimensions[column_letter].width = 15  # 设置列宽为15

            # 设置行背景颜色
            for row in range(2, len(df_logs) + 2):  # 从第二行开始
                phase = df_logs.iloc[row - 2]["阶段"]
                if phase == "新一轮放娃":
                    # 设置整行背景颜色为浅灰色
                    for col in range(1, len(df_logs.columns) + 1):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color='EDEDED', end_color='EDEDED',
                                                                        fill_type='solid')
                elif phase == "初始化":
                    # 设置整行背景颜色为中灰色
                    for col in range(1, len(df_logs.columns) + 1):
                        ws.cell(row=row, column=col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3',
                                                                        fill_type='solid')

            # 保存修改
            wb.save(tmp.name)

            tmp_path = tmp.name

        return tmp_path, df_results

    except Exception as e:
        return f"错误发生：{str(e)}", None


# Gradio界面
with gr.Blocks(title="娃娃宝塔模拟器") as demo:
    gr.Markdown("## 🏯 娃娃宝塔游戏模拟器")

    with gr.Row():
        with gr.Column(scale=2):
            # 颜色配置水平布局
            with gr.Row():
                doll_config_red = gr.Number(label="红色", value=100, minimum=0, step=1, min_width=80)
                doll_config_orange = gr.Number(label="橙色", value=100, minimum=0, step=1, min_width=80)
                doll_config_yellow = gr.Number(label="黄色", value=100, minimum=0, step=1, min_width=80)
                doll_config_green = gr.Number(label="绿色", value=100, minimum=0, step=1, min_width=80)
                doll_config_blue = gr.Number(label="蓝色", value=100, minimum=0, step=1, min_width=80)
                doll_config_purple = gr.Number(label="紫色", value=100, minimum=0, step=1, min_width=80)
                doll_config_pink = gr.Number(label="粉色", value=100, minimum=0, step=1, min_width=80)
                doll_config_black = gr.Number(label="黑色", value=100, minimum=0, step=1, min_width=80)
                doll_config_white = gr.Number(label="白色", value=100, minimum=0, step=1, min_width=80)
                doll_config_brown = gr.Number(label="棕色", value=100, minimum=0, step=1, min_width=80)
                doll_config_gray = gr.Number(label="灰色", value=100, minimum=0, step=1, min_width=80)
                doll_config_gold = gr.Number(label="金色", value=100, minimum=0, step=1, min_width=80)
                doll_config_silver = gr.Number(label="银色", value=100, minimum=0, step=1, min_width=80)
            wish_colors = gr.CheckboxGroup(
                label="选择许愿色（可多选）",
                choices=DEFAULT_COLORS
            )
            with gr.Row():
                milk1 = gr.Number(label="第一次奶", value=0, minimum=0, step=1, min_width=80)
                milk2 = gr.Number(label="第二次奶", value=0, minimum=0, step=1, min_width=80)
                milk3 = gr.Number(label="第三次奶", value=0, minimum=0, step=1, min_width=80)
            initial_draw = gr.Number(
                label="初始抽取数量",
                value=9,
                minimum=1,
                step=1
            )
            doll_exchange = gr.Number(
                label="娃娃兑换率（N娃娃=1礼包）",
                value=18,
                minimum=1,
                step=1
            )
            total_games = gr.Number(
                label="模拟总局数",
                value=100,
                minimum=1,
                step=1
            )
            max_rounds = gr.Number(
                label="最大回合数",
                value=100,
                minimum=1,
                step=1,
                visible=False
            )
            btn_run = gr.Button("开始模拟", variant="primary")

        with gr.Column(scale=3):
            file_output = gr.File(label="下载结果文件")
            gr.Markdown("### 最近10局结果")
            result_table = gr.Dataframe(
                headers=["游戏ID", "收获娃娃", "大礼包", "回合数"],
                datatype=["number", "number", "number", "number"]
            )

    # 运行模拟
    btn_run.click(
        run_simulation,
        inputs=[doll_config_red, doll_config_orange, doll_config_yellow, doll_config_green, doll_config_blue,
                doll_config_purple, doll_config_pink, doll_config_black, doll_config_white, doll_config_brown,
                doll_config_gray, doll_config_gold, doll_config_silver, wish_colors, initial_draw, doll_exchange,
                total_games, max_rounds, milk1, milk2, milk3],
        outputs=[file_output, result_table]
    )

if __name__ == "__main__":
    demo.launch()

    demo.app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],  # Allow all origins
        allow_credentials=True,
        allow_methods=["*"],  # Allow all methods
        allow_headers=["*"],  # Allow all headers
    )
